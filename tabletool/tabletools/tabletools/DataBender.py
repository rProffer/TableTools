'''
This class gathers data from excel in a 2d array using openpyxl
Author: Robyn Proffer July 2018
'''


'''TODO

'''

#used libraries and modules
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl import load_workbook



class DataBender:

    def Retrieve(self, file, mini = 0, maxi = None, colbreak = True):
        """
        Retrieve data from within a workbook.
        File is the name of the workbook to pull from, mini is the minimum row which defaults to the topmost row, maxi
        is the final row to process and defaults to the final row, colbreak is a boolean which determines
        whether or not the processing of a row ends on a blank column, which defaults to yes.
        """
        sheet = []
        inBook = openpyxl.load_workbook(filename = file)
        dataSheet = inBook.active #could do getsheet here too
        count = 0
        if maxi == None:
            maxi = dataSheet.max_row
        header = list(dataSheet.rows)[0]
        for row in dataSheet.iter_rows(min_row = mini, max_row = maxi):
            newrow = []
            for col in row:
                if col.value == None:
                    if colbreak == True:
                        break
                newrow.append(str(col.value))
            sheet.append(newrow)
        inBook.close
        return sheet
    
    def putOut(self, data, title, heading): #title need be a filename prepped for saving. EG: terminal.xlsx. data must be a two dimensional array
        outBook = openpyxl.Workbook()
        outSheet = outBook.active        
        outSheet.append(heading) #heading needs to be a 1d array         
        for a in data: #figure out the structure underlying data
            outSheet.append(a)
        outBook.save(title)
        outBook.close
